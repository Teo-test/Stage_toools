#!/usr/bin/env python3
"""
Excel Merger — Fusion de tous les fichiers Excel d'un dossier en un seul classeur
Lit tous les .xlsx / .xls / .csv d'un répertoire et les regroupe en onglets.

Usage: python excel_merger.py [dossier] [-o fichier_sortie.xlsx]
"""

import sys
import os
import re
import argparse
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError as e:
    print(f"[ERREUR] Dépendance manquante : {e}")
    print("Installe : pip install pandas openpyxl xlrd")
    sys.exit(1)

# ─── Couleurs terminal ────────────────────────────────────────────────────────
class C:
    BOLD    = "\033[1m"
    DIM     = "\033[2m"
    CYAN    = "\033[96m"
    GREEN   = "\033[92m"
    YELLOW  = "\033[93m"
    RED     = "\033[91m"
    BLUE    = "\033[94m"
    MAGENTA = "\033[95m"
    RESET   = "\033[0m"

def titre(texte):
    largeur = 60
    print(f"\n{C.CYAN}{C.BOLD}{'─' * largeur}{C.RESET}")
    print(f"{C.CYAN}{C.BOLD}  {texte}{C.RESET}")
    print(f"{C.CYAN}{C.BOLD}{'─' * largeur}{C.RESET}")

def info(texte):   print(f"  {C.BLUE}ℹ {C.RESET}{texte}")
def ok(texte):     print(f"  {C.GREEN}✔ {C.RESET}{texte}")
def warn(texte):   print(f"  {C.YELLOW}⚠ {C.RESET}{texte}")
def erreur(texte): print(f"  {C.RED}✘ {C.RESET}{texte}")

def menu_numerote(titre_menu, options, allow_multiple=False):
    print(f"\n{C.BOLD}  {titre_menu}{C.RESET}")
    for i, opt in enumerate(options, 1):
        print(f"    {C.CYAN}{i:>2}{C.RESET}. {opt}")
    print()
    while True:
        if allow_multiple:
            entree = input(f"  {C.BOLD}Choix (ex: 1 3 5 ou 'all'){C.RESET} : ").strip()
            if entree.lower() == "all":
                return list(range(len(options)))
            try:
                choix = [int(x) - 1 for x in entree.split()]
                if all(0 <= c < len(options) for c in choix) and choix:
                    return choix
            except ValueError:
                pass
        else:
            entree = input(f"  {C.BOLD}Choix{C.RESET} : ").strip()
            try:
                c = int(entree) - 1
                if 0 <= c < len(options):
                    return c
            except ValueError:
                pass
        erreur("Entrée invalide, réessaie.")

def input_prompt(texte, defaut=None):
    if defaut:
        rep = input(f"  {C.BOLD}{texte}{C.RESET} [{C.DIM}{defaut}{C.RESET}] : ").strip()
        return rep if rep else defaut
    return input(f"  {C.BOLD}{texte}{C.RESET} : ").strip()

def slugify(texte, max_len=31):
    """
    Nettoie un nom pour en faire un nom d'onglet Excel valide (max 31 car.).
    Caractères interdits dans les onglets : \ / * ? : [ ]
    """
    texte = re.sub(r'[\\/*?:\[\]]', '', str(texte))
    texte = texte.strip()
    return texte[:max_len] if texte else "Feuille"

def nom_unique(nom, existants):
    """Garantit un nom d'onglet unique en ajoutant un suffixe si nécessaire."""
    if nom not in existants:
        return nom
    base = nom[:28]
    i = 2
    while f"{base}_{i}" in existants:
        i += 1
    return f"{base}_{i}"

# ─── Extensions prises en charge ─────────────────────────────────────────────
EXTENSIONS = {'.xlsx', '.xls', '.xlsm', '.csv', '.tsv'}

# ─── Scan du dossier ─────────────────────────────────────────────────────────

def scanner_dossier(dossier, recursif=False):
    """
    Liste tous les fichiers Excel/CSV du dossier.
    Retourne une liste de Path triée par nom.
    """
    p = Path(dossier)
    if not p.is_dir():
        erreur(f"Dossier introuvable : {dossier}")
        return []

    if recursif:
        fichiers = [f for f in p.rglob('*') if f.suffix.lower() in EXTENSIONS and f.is_file()]
    else:
        fichiers = [f for f in p.iterdir() if f.suffix.lower() in EXTENSIONS and f.is_file()]

    return sorted(fichiers, key=lambda f: f.name.lower())

# ─── Lecture des fichiers ─────────────────────────────────────────────────────

def lire_fichier(chemin: Path):
    """
    Lit un fichier Excel ou CSV.
    Retourne un dict {nom_onglet: DataFrame} — un seul onglet pour CSV,
    plusieurs pour les fichiers Excel multi-feuilles.
    """
    ext = chemin.suffix.lower()
    onglets = {}

    try:
        if ext in ('.csv', '.tsv'):
            sep = '\t' if ext == '.tsv' else ','
            df = pd.read_csv(chemin, sep=sep, encoding='utf-8-sig', low_memory=False)
            onglets[chemin.stem] = df

        elif ext in ('.xlsx', '.xlsm'):
            sheets = pd.read_excel(chemin, sheet_name=None, engine='openpyxl')
            for nom, df in sheets.items():
                cle = f"{chemin.stem}_{nom}" if len(sheets) > 1 else chemin.stem
                onglets[cle] = df

        elif ext == '.xls':
            try:
                sheets = pd.read_excel(chemin, sheet_name=None, engine='xlrd')
            except Exception:
                sheets = pd.read_excel(chemin, sheet_name=None)
            for nom, df in sheets.items():
                cle = f"{chemin.stem}_{nom}" if len(sheets) > 1 else chemin.stem
                onglets[cle] = df

    except Exception as e:
        erreur(f"Impossible de lire {chemin.name} : {e}")
        return {}

    return onglets

def charger_tous(fichiers):
    """
    Charge tous les fichiers et construit la liste des onglets à fusionner.
    Retourne une liste de dicts avec métadonnées.
    """
    onglets_list = []
    idx = 0
    for chemin in fichiers:
        info(f"Lecture de {C.BOLD}{chemin.name}{C.RESET} …")
        onglets = lire_fichier(chemin)
        if not onglets:
            warn(f"  Aucune donnée lue dans {chemin.name}")
            continue
        for nom_onglet, df in onglets.items():
            idx += 1
            onglets_list.append({
                "idx":      idx,
                "fichier":  chemin.name,
                "origine":  str(chemin),
                "onglet":   nom_onglet,
                "df":       df,
                "lignes":   len(df),
                "colonnes": len(df.columns),
                "vide":     df.empty,
            })
            statut = f"{C.YELLOW}[vide]{C.RESET}" if df.empty else f"{len(df)} lignes × {len(df.columns)} col."
            ok(f"  Onglet {C.BOLD}{nom_onglet}{C.RESET} — {statut}")
    return onglets_list

# ─── Affichage ────────────────────────────────────────────────────────────────

def afficher_resume(onglets_list):
    titre("ONGLETS TROUVÉS")
    if not onglets_list:
        warn("Aucune donnée chargée.")
        return
    print(f"  {'N°':>3}  {'Fichier':<30}  {'Onglet':<25}  {'Lignes':>6}  {'Cols':>4}")
    print(f"  {'─'*3}  {'─'*30}  {'─'*25}  {'─'*6}  {'─'*4}")
    for o in onglets_list:
        vide   = f" {C.YELLOW}[vide]{C.RESET}" if o['vide'] else ""
        fic    = o['fichier'][:29]
        ong    = o['onglet'][:24]
        print(f"  {C.CYAN}{o['idx']:>3}{C.RESET}  "
              f"  {fic:<29}  "
              f"  {C.BOLD}{ong:<24}{C.RESET}  "
              f"  {o['lignes']:>5}  "
              f"  {o['colonnes']:>3}  {vide}")

def afficher_detail(onglets_list):
    titre("DÉTAIL D'UN ONGLET")
    idx = menu_numerote(
        "Quel onglet ?",
        [f"[{o['fichier']}] {o['onglet']} — {o['lignes']} lignes" for o in onglets_list]
    )
    o = onglets_list[idx]
    print(f"\n  {C.BOLD}{C.MAGENTA}{o['onglet']}{C.RESET}")
    print(f"  Fichier source : {o['fichier']}")
    print(f"  Lignes         : {o['lignes']}")
    print(f"  Colonnes       : {o['colonnes']}")
    if not o['df'].empty:
        print(f"  Colonnes noms  : {', '.join(str(c) for c in o['df'].columns)}")
        print(f"\n{o['df'].head(10).to_string(index=False)}\n")
    else:
        warn("  Aucune donnée dans cet onglet.")

# ─── Styles Excel ─────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="2F5496")   # Bleu foncé
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT     = Font(name="Arial", size=10)
CELL_ALIGN    = Alignment(vertical="top")

ALT_FILL_1    = PatternFill("solid", start_color="EEF2FA")   # Bleu très clair (lignes paires)
ALT_FILL_2    = PatternFill("solid", start_color="FFFFFF")   # Blanc (lignes impaires)

BORDER_THIN   = Border(
    left   = Side(style='thin', color='C9C9C9'),
    right  = Side(style='thin', color='C9C9C9'),
    top    = Side(style='thin', color='C9C9C9'),
    bottom = Side(style='thin', color='C9C9C9'),
)

def ajuster_largeur_colonnes(ws, df):
    """Ajuste automatiquement la largeur des colonnes selon leur contenu."""
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if not df.empty else 0
        )
        # Borne : min 8, max 60
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 8), 60)

def ecrire_onglet(wb, nom_onglet, df, avec_style=True, figer_en_tete=True, avec_filtre=True):
    """
    Crée un onglet dans le classeur et y écrit le DataFrame avec formatage.
    """
    ws = wb.create_sheet(title=nom_onglet)

    if df.empty:
        ws['A1'] = "(onglet vide)"
        ws['A1'].font = Font(name="Arial", italic=True, color="999999")
        return ws

    # ── En-têtes ──────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        if avec_style:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border    = BORDER_THIN
    ws.row_dimensions[1].height = 22

    # ── Données ───────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for col_idx, valeur in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valeur)
            if avec_style:
                cell.font      = CELL_FONT
                cell.alignment = CELL_ALIGN
                cell.border    = BORDER_THIN
                cell.fill      = ALT_FILL_1 if row_idx % 2 == 0 else ALT_FILL_2

    # ── Options ───────────────────────────────────────────────────────────────
    if figer_en_tete:
        ws.freeze_panes = "A2"

    if avec_filtre:
        last_col = get_column_letter(len(df.columns))
        last_row = len(df) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    if avec_style:
        ajuster_largeur_colonnes(ws, df)

    return ws

def ecrire_sommaire(wb, onglets_list, noms_onglets):
    """
    Crée une feuille 'Sommaire' listant tous les onglets fusionnés
    avec hyperliens vers chacun.
    """
    ws = wb.create_sheet(title="Sommaire", index=0)
    ws.sheet_view.showGridLines = False

    # Titre principal
    ws.merge_cells("A1:E1")
    cell_titre = ws["A1"]
    cell_titre.value     = "📊 Classeur fusionné — Sommaire"
    cell_titre.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    cell_titre.fill      = PatternFill("solid", start_color="1F3864")
    cell_titre.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Sous-titre
    ws.merge_cells("A2:E2")
    ws["A2"].value     = f"{len(onglets_list)} onglet(s) importé(s)"
    ws["A2"].font      = Font(name="Arial", italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # En-têtes du tableau
    entetes = ["N°", "Fichier source", "Onglet d'origine", "Lignes", "Colonnes"]
    for col_idx, h in enumerate(entetes, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = BORDER_THIN
    ws.row_dimensions[4].height = 20

    # Lignes de données avec hyperliens
    for i, (o, nom_onglet) in enumerate(zip(onglets_list, noms_onglets), start=5):
        donnees = [o['idx'], o['fichier'], o['onglet'], o['lignes'], o['colonnes']]
        for col_idx, val in enumerate(donnees, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.font      = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border    = BORDER_THIN
            cell.fill      = ALT_FILL_1 if i % 2 == 0 else ALT_FILL_2

        # Hyperlien sur le nom de l'onglet
        lien_cell = ws.cell(row=i, column=2)
        lien_cell.hyperlink = f"#{nom_onglet}!A1"
        lien_cell.font      = Font(name="Arial", size=10, color="1F4E79", underline="single")

    # Largeurs colonnes du sommaire
    for col, largeur in zip(['A','B','C','D','E'], [6, 35, 30, 10, 10]):
        ws.column_dimensions[col].width = largeur

    ws.freeze_panes = "A5"
    return ws

# ─── Fusion ───────────────────────────────────────────────────────────────────

def fusionner(onglets_list, selection, chemin_sortie,
              avec_style=True, figer_en_tete=True, avec_filtre=True,
              avec_sommaire=True):
    """
    Crée le classeur final en écrivant chaque onglet sélectionné.
    """
    titre("FUSION EN COURS")
    wb = openpyxl.Workbook()
    # Supprimer la feuille par défaut
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    noms_existants = set()
    noms_onglets   = []
    onglets_traites = []

    for idx in selection:
        o   = onglets_list[idx]
        nom = slugify(o['onglet'])
        nom = nom_unique(nom, noms_existants)
        noms_existants.add(nom)
        noms_onglets.append(nom)
        onglets_traites.append(o)

        info(f"Écriture de {C.BOLD}{nom}{C.RESET} ({o['lignes']} lignes)…")
        ecrire_onglet(wb, nom, o['df'],
                      avec_style=avec_style,
                      figer_en_tete=figer_en_tete,
                      avec_filtre=avec_filtre)
        ok(f"  Onglet {C.BOLD}{nom}{C.RESET} écrit.")

    if avec_sommaire and onglets_traites:
        ecrire_sommaire(wb, onglets_traites, noms_onglets)
        ok(f"  Onglet {C.BOLD}Sommaire{C.RESET} créé.")

    sortie = Path(chemin_sortie)
    sortie.parent.mkdir(parents=True, exist_ok=True)
    wb.save(sortie)
    ok(f"Fichier sauvegardé : {C.BOLD}{sortie}{C.RESET}")
    return sortie

# ─── Menus thématiques ────────────────────────────────────────────────────────

def choisir_onglets(onglets_list, message="Quels onglets fusionner ?"):
    labels = [f"[{o['fichier']}]  {o['onglet']:<25}  {o['lignes']} lignes" for o in onglets_list]
    idxs   = menu_numerote(message, labels + ["Tous"], allow_multiple=True)
    if len(onglets_list) in idxs:
        return list(range(len(onglets_list)))
    return idxs

def menu_fusionner(onglets_list, chemin_sortie):
    titre("OPTIONS DE FUSION")

    selection = choisir_onglets(onglets_list, "Quels onglets inclure dans le fichier final ?")

    style_idx = menu_numerote(
        "Style des cellules",
        ["Avec mise en forme (couleurs, bordures, tailles)", "Sans mise en forme (données brutes)"]
    )
    avec_style = (style_idx == 0)

    options_idx = menu_numerote(
        "Options supplémentaires",
        ["Tout activer (en-tête figé + filtres + sommaire)",
         "En-tête figé uniquement",
         "Filtres automatiques uniquement",
         "Aucune option (données brutes)"]
    )
    figer_en_tete = options_idx in (0, 1)
    avec_filtre   = options_idx in (0, 2)
    avec_sommaire = (options_idx == 0)

    fusionner(onglets_list, selection, chemin_sortie,
              avec_style=avec_style,
              figer_en_tete=figer_en_tete,
              avec_filtre=avec_filtre,
              avec_sommaire=avec_sommaire)

def menu_fusion_rapide(onglets_list, chemin_sortie):
    """Fusion de tous les onglets avec toutes les options en un clic."""
    titre("FUSION RAPIDE (tout inclus)")
    non_vides = [i for i, o in enumerate(onglets_list) if not o['vide']]
    if not non_vides:
        warn("Tous les onglets sont vides.")
        return
    info(f"{len(non_vides)} onglet(s) non vide(s) → {C.BOLD}{chemin_sortie}{C.RESET}")
    fusionner(onglets_list, non_vides, chemin_sortie,
              avec_style=True, figer_en_tete=True,
              avec_filtre=True, avec_sommaire=True)

def menu_changer_sortie(chemin_actuel):
    nouveau = input_prompt("Nouveau chemin de sortie", chemin_actuel)
    if not nouveau.endswith('.xlsx'):
        nouveau += '.xlsx'
    ok(f"Fichier de sortie → {C.BOLD}{nouveau}{C.RESET}")
    return nouveau

# ─── Menu principal ───────────────────────────────────────────────────────────

MENU_PRINCIPAL = [
    "Fusion rapide (tous les onglets, toutes les options)",
    "Fusion personnalisée (choisir onglets + options)",
    "Voir le détail d'un onglet",
    "Résumé des onglets chargés",
    "Changer le fichier de sortie",
    "Charger un autre dossier",
    "Quitter",
]

def charger_dossier(dossier, recursif=False):
    fichiers = scanner_dossier(dossier, recursif=recursif)
    if not fichiers:
        warn(f"Aucun fichier Excel/CSV trouvé dans : {dossier}")
        return None, []
    info(f"{len(fichiers)} fichier(s) trouvé(s) :")
    for f in fichiers:
        print(f"    {C.DIM}•{C.RESET} {f.name}")
    onglets_list = charger_tous(fichiers)
    return fichiers, onglets_list

def main():
    parser = argparse.ArgumentParser(
        description="Excel Merger — Fusionne tous les .xlsx/.csv d'un dossier en un seul classeur")
    parser.add_argument("dossier", nargs="?", default=".",
                        help="Dossier contenant les fichiers Excel (défaut: dossier courant)")
    parser.add_argument("-o", "--output", default="fusion_excel.xlsx",
                        help="Fichier de sortie (défaut: fusion_excel.xlsx)")
    parser.add_argument("-r", "--recursif", action="store_true",
                        help="Parcourir les sous-dossiers récursivement")
    args = parser.parse_args()

    print(f"\n{C.CYAN}{C.BOLD}")
    print("  ╔══════════════════════════════════════════╗")
    print("  ║       EXCEL MERGER  v1.0                 ║")
    print("  ║  Fusion multi-fichiers → un classeur     ║")
    print("  ╚══════════════════════════════════════════╝")
    print(C.RESET)

    onglets_list  = []
    dossier_actif = None
    chemin_sortie = args.output
    if not chemin_sortie.endswith('.xlsx'):
        chemin_sortie += '.xlsx'

    if args.dossier:
        fichiers, onglets_list = charger_dossier(args.dossier, recursif=args.recursif)
        if onglets_list:
            dossier_actif = args.dossier
            afficher_resume(onglets_list)
        else:
            info("Aucun onglet chargé. Spécifie un autre dossier via l'option 6.")

    while True:
        titre("MENU PRINCIPAL")
        if dossier_actif:
            info(f"Dossier  : {C.BOLD}{Path(dossier_actif).resolve()}{C.RESET}  "
                 f"({len(onglets_list)} onglet(s))")
            info(f"Sortie   : {C.BOLD}{chemin_sortie}{C.RESET}")
        else:
            warn("Aucun dossier chargé")

        choix = menu_numerote("Que veux-tu faire ?", MENU_PRINCIPAL)

        if not onglets_list and choix not in (4, 5, 6):
            warn("Charge d'abord un dossier (option 6).")
            continue

        if   choix == 0: menu_fusion_rapide(onglets_list, chemin_sortie)
        elif choix == 1: menu_fusionner(onglets_list, chemin_sortie)
        elif choix == 2: afficher_detail(onglets_list)
        elif choix == 3: afficher_resume(onglets_list)
        elif choix == 4:
            chemin_sortie = menu_changer_sortie(chemin_sortie)
        elif choix == 5:
            dossier = input_prompt("Chemin du dossier à analyser", ".").strip('"').strip("'")
            recursif = input_prompt("Inclure les sous-dossiers ? (o/n)", "n").lower() == 'o'
            fichiers, onglets_list = charger_dossier(dossier, recursif=recursif)
            if onglets_list:
                dossier_actif = dossier
                afficher_resume(onglets_list)
        elif choix == 6:
            print(f"\n  {C.GREEN}Au revoir !{C.RESET}\n")
            break

if __name__ == "__main__":
    main()
